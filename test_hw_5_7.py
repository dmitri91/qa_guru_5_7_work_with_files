import os.path
import requests
import csv
import time
import xlrd
from selenium import webdriver
from selene import browser
from openpyxl import load_workbook
from pypdf import PdfReader
from zipfile import ZipFile, ZIP_DEFLATED


def test_downloaded_file_size():
    # TODO сохранять и читать из tmp, использовать универсальный путь
    url = 'https://selenium.dev/images/selenium_logo_square_green.png'

    r = requests.get(url)
    dir_root_path = os.path.dirname(os.path.abspath(__file__))
    path_tmp = os.path.join(dir_root_path, 'tmp', 'selenium_logo.png')

    with open(path_tmp, 'wb') as file:
        file.write(r.content)
        size = os.path.getsize(path_tmp)

    assert size == 30803


def test_download_file_browser():
    dir_root_path = os.path.dirname(os.path.abspath(__file__))
    path_tmp = os.path.join(dir_root_path, 'tmp')

    options = webdriver.ChromeOptions()
    prefs = {
        "download.default_directory": path_tmp,
        "download.prompt_for_download": False
    }
    options.add_experimental_option("prefs", prefs)

    browser.config.driver_options = options

    browser.open("https://github.com/pytest-dev/pytest")
    browser.element(".d-none .Button-label").click()
    browser.element('[data-open-app="link"]').click()
    time.sleep(3)

    pytest_dir = os.path.join(dir_root_path, 'tmp', 'pytest-main.zip')
    size = os.path.getsize(pytest_dir)
    assert size == 1564394


def test_work_with_csv():
    dir_root_path = os.path.dirname(os.path.abspath(__file__))
    path_file_1 = os.path.join(dir_root_path, 'tmp', 'file_1.csv')
    # TODO оформить в тест, добавить ассерты и использовать универсальный путь
    with open(path_file_1, 'w', newline='') as csvfile:
        csvwriter = csv.writer(csvfile, delimiter=',')
        csvwriter.writerow(['Anna', 'Pavel', 'Peter'])
        csvwriter.writerow(['Alex', 'Serj', 'Yana'])
    with open(path_file_1) as csvfile:
        csvreader = csv.reader(csvfile)
        list_name = []
        for row in csvreader:
            print(row)
            list_name.append(row)
        assert list_name[0] == ['Anna', 'Pavel', 'Peter']
        assert list_name[1] == ['Alex', 'Serj', 'Yana']


def test_work_with_xls():
    # TODO оформить в тест, добавить ассерты и использовать универсальный путь
    dir_root_path = os.path.dirname(os.path.abspath(__file__))
    path_xls = os.path.join(dir_root_path, 'resources', 'file_example_XLS_10.xls')
    book = xlrd.open_workbook(path_xls)
    print(f'Количество листов {book.nsheets}')
    print(f'Имена листов {book.sheet_names()}')
    sheet = book.sheet_by_index(0)
    print(f'Количество столбцов {sheet.ncols}')
    print(f'Количество строк {sheet.nrows}')
    print(f'Пересечение строки 9 и столбца 1 = {sheet.cell_value(rowx=9, colx=1)}')
    # печать всех строк по очереди
    for rx in range(sheet.nrows):
        print(sheet.row(rx))

    assert book.nsheets == 1
    assert sheet.ncols == 8
    assert sheet.cell_value(rowx=9, colx=1) == 'Vincenza'


def test_work_with_xlsx():
    # TODO оформить в тест, добавить ассерты и использовать универсальный путь
    dir_root_path = os.path.dirname(os.path.abspath(__file__))
    path_xlsx = os.path.join(dir_root_path, 'resources', 'file_example_XLSX_50.xlsx')

    workbook = load_workbook(path_xlsx)
    sheet = workbook.active
    value = sheet.cell(row=3, column=2).value
    print(value)
    assert value == 'Mara'


def test_work_with_pdf():
    # TODO оформить в тест, добавить ассерты и использовать универсальный путь
    dir_root_path = os.path.dirname(os.path.abspath(__file__))
    path_xlsx = os.path.join(dir_root_path, 'resources', 'docs-pytest-org-en-latest.pdf')
    reader = PdfReader(path_xlsx)
    number_of_pages = len(reader.pages)
    page = reader.pages[0]
    text = page.extract_text()
    print(page)
    print(number_of_pages)
    print(text)
    assert number_of_pages == 412


def test_work_with_zip():
    dir_root_path = os.path.dirname(os.path.abspath(__file__))
    path_resources = os.path.join(dir_root_path, 'resources')
    list_dir = os.listdir(path_resources)
    with ZipFile('test.zip', mode='w', compression=ZIP_DEFLATED) as zf:
        for file in list_dir:
            add_file = os.path.join(path_resources, file)
            zf.write(add_file)
    list_size = {'docs-pytest-org-en-latest.pdf': 1739253,
                 'file_example_XLSX_50.xlsx': 7360,
                 'file_example_XLS_10.xls': 8704,
                 'hello.zip': 128}

    with ZipFile('test.zip', mode='a') as zf:
        for file in zf.infolist():
            size = file.file_size
            file_name = os.path.basename(file.filename)
            assert (file_name, size) in list_size.items()
